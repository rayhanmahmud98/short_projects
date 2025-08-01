import pygame
import random
import sys
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# Initialize pygame
pygame.init()

# Game constants
WIDTH, HEIGHT = 400, 600
FPS = 60
GRAVITY = 0.25
BIRD_JUMP = -5
PIPE_SPEED = 3
PIPE_GAP = 150
PIPE_FREQUENCY = 1500  # milliseconds

# Colors
WHITE = (255, 255, 255)
BLACK = (0, 0, 0)
GREEN = (0, 255, 0)
SKY_BLUE = (135, 206, 235)
RED = (255, 0, 0)

# Set up the display
screen = pygame.display.set_mode((WIDTH, HEIGHT))
pygame.display.set_caption("Floppy Bird")
clock = pygame.time.Clock()

# Game variables
score = 0
high_score = 0
game_active = False
player_name = ""
name_input_active = True
font = pygame.font.SysFont('Arial', 30)
small_font = pygame.font.SysFont('Arial', 20)

# Excel file setup
EXCEL_FILE = "floppy_bird_scores.xlsx"
SHEET_NAME = "High Scores"

class Bird:
    def __init__(self):
        self.x = 100
        self.y = HEIGHT // 2
        self.velocity = 0
        self.radius = 15
        
    def update(self):
        # Apply gravity
        self.velocity += GRAVITY
        self.y += self.velocity
        
        # Keep bird on screen
        if self.y > HEIGHT - self.radius:
            self.y = HEIGHT - self.radius
            self.velocity = 0
        if self.y < self.radius:
            self.y = self.radius
            self.velocity = 0
            
    def jump(self):
        self.velocity = BIRD_JUMP
        
    def draw(self):
        pygame.draw.circle(screen, (255, 255, 0), (self.x, int(self.y)), self.radius)
        # Draw eye
        pygame.draw.circle(screen, BLACK, (self.x + 5, int(self.y) - 3), 3)
        # Draw beak
        pygame.draw.polygon(screen, (255, 165, 0), 
                           [(self.x + 15, int(self.y)), 
                            (self.x + 25, int(self.y)), 
                            (self.x + 15, int(self.y) + 5)])
        
    def get_mask(self):
        return pygame.Rect(self.x - self.radius, self.y - self.radius, 
                          self.radius * 2, self.radius * 2)

class Pipe:
    def __init__(self):
        self.x = WIDTH
        self.height = random.randint(100, HEIGHT - 200)
        self.top_pipe = pygame.Rect(self.x, 0, 50, self.height)
        self.bottom_pipe = pygame.Rect(self.x, self.height + PIPE_GAP, 50, HEIGHT - self.height - PIPE_GAP)
        self.passed = False
        
    def update(self):
        self.x -= PIPE_SPEED
        self.top_pipe.x = self.x
        self.bottom_pipe.x = self.x
        
    def draw(self):
        pygame.draw.rect(screen, GREEN, self.top_pipe)
        pygame.draw.rect(screen, GREEN, self.bottom_pipe)
        
    def collide(self, bird):
        bird_rect = bird.get_mask()
        return bird_rect.colliderect(self.top_pipe) or bird_rect.colliderect(self.bottom_pipe)

def draw_score():
    score_surface = font.render(f"Score: {score}", True, BLACK)
    screen.blit(score_surface, (10, 10))
    
    high_score_surface = font.render(f"High: {high_score}", True, BLACK)
    screen.blit(high_score_surface, (10, 50))

def draw_game_over():
    game_over_surface = font.render("Game Over!", True, RED)
    restart_surface = small_font.render("Press SPACE to play again", True, BLACK)
    
    screen.blit(game_over_surface, (WIDTH // 2 - game_over_surface.get_width() // 2, HEIGHT // 2 - 50))
    screen.blit(restart_surface, (WIDTH // 2 - restart_surface.get_width() // 2, HEIGHT // 2 + 50))

def draw_name_input():
    input_title = font.render("Enter your name:", True, BLACK)
    name_surface = font.render(player_name, True, BLACK)
    start_surface = small_font.render("Press ENTER to start", True, BLACK)
    
    screen.blit(input_title, (WIDTH // 2 - input_title.get_width() // 2, HEIGHT // 2 - 50))
    pygame.draw.rect(screen, BLACK, (WIDTH // 2 - 100, HEIGHT // 2, 200, 40), 2)
    screen.blit(name_surface, (WIDTH // 2 - name_surface.get_width() // 2, HEIGHT // 2 + 5))
    screen.blit(start_surface, (WIDTH // 2 - start_surface.get_width() // 2, HEIGHT // 2 + 60))

def reset_game():
    global score, game_active
    bird.y = HEIGHT // 2
    bird.velocity = 0
    pipes.clear()
    score = 0
    game_active = True
    last_pipe = pygame.time.get_ticks()

def initialize_excel_file():
    """Create or load the Excel file with proper headers"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["Player Name", "Score", "Date"])
        # Style the headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
        wb.save(EXCEL_FILE)

def save_score_to_excel():
    """Save the score to Excel, keeping only the highest score per player"""
    initialize_excel_file()
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    
    # Create dictionary of existing scores {name: (score, row_num)}
    existing_scores = {}
    for row in range(2, ws.max_row + 1):
        name = ws[f"A{row}"].value
        score_val = ws[f"B{row}"].value
        if name and score_val:
            existing_scores[name] = (int(score_val), row)
    
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    if player_name in existing_scores:
        # Player exists - update only if current score is higher
        existing_score, row_num = existing_scores[player_name]
        if score > existing_score:
            ws[f"B{row_num}"] = score
            ws[f"C{row_num}"] = current_date
    else:
        # New player - add new row
        ws.append([player_name, score, current_date])
    
    # Sort by score (descending)
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Only include rows with player names
            data.append(row)
    
    # Sort by score (column B) descending
    data.sort(key=lambda x: x[1], reverse=True)
    
    # Clear existing data (except headers)
    ws.delete_rows(2, ws.max_row)
    
    # Write sorted data back
    for row in data:
        ws.append(row)
    
    wb.save(EXCEL_FILE)

def show_high_scores():
    """Display top 5 high scores from Excel file"""
    if not os.path.exists(EXCEL_FILE):
        return
    
    screen.fill(SKY_BLUE)
    title = font.render("High Scores", True, BLACK)
    screen.blit(title, (WIDTH // 2 - title.get_width() // 2, 50))
    
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[SHEET_NAME]
        
        y_pos = 120
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), 1):
            if row[0]:  # If player name exists
                name, scr, date = row[0], row[1], row[2]
                score_text = small_font.render(f"{i}. {name}: {scr} ({date})", True, BLACK)
                screen.blit(score_text, (WIDTH // 2 - score_text.get_width() // 2, y_pos))
                y_pos += 40
    except Exception as e:
        print(f"Error reading Excel file: {e}")
    
    back_text = small_font.render("Press SPACE to continue", True, BLACK)
    screen.blit(back_text, (WIDTH // 2 - back_text.get_width() // 2, HEIGHT - 50))
    pygame.display.update()
    
    waiting = True
    while waiting:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                sys.exit()
            if event.type == pygame.KEYDOWN and event.key == pygame.K_SPACE:
                waiting = False

# Game objects
bird = Bird()
pipes = []
last_pipe = pygame.time.get_ticks()

# Initialize Excel file
initialize_excel_file()

# Main game loop
running = True
while running:
    clock.tick(FPS)
    
    # Event handling
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False
        
        if name_input_active:
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_RETURN and player_name:
                    name_input_active = False
                    show_high_scores()
                    reset_game()
                elif event.key == pygame.K_BACKSPACE:
                    player_name = player_name[:-1]
                else:
                    # Only allow letters and numbers, limit length
                    if len(player_name) < 12 and event.unicode.isalnum():
                        player_name += event.unicode
        else:
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_SPACE and game_active:
                    bird.jump()
                if event.key == pygame.K_SPACE and not game_active:
                    if player_name:  # Only reset if we have a player name
                        save_score_to_excel()
                        show_high_scores()
                        reset_game()
    
    # Fill background
    screen.fill(SKY_BLUE)
    
    if name_input_active:
        draw_name_input()
    elif game_active:
        # Bird update
        bird.update()
        
        # Pipe generation
        time_now = pygame.time.get_ticks()
        if time_now - last_pipe > PIPE_FREQUENCY:
            pipes.append(Pipe())
            last_pipe = time_now
            
        # Pipe update and collision
        for pipe in pipes[:]:
            pipe.update()
            
            # Check for collision
            if pipe.collide(bird):
                game_active = False
                save_score_to_excel()
                
            # Check if pipe passed
            if not pipe.passed and pipe.x < bird.x:
                pipe.passed = True
                score += 1
                if score > high_score:
                    high_score = score
                    
            # Remove off-screen pipes
            if pipe.x < -50:
                pipes.remove(pipe)
                
        # Check if bird hit the ground
        if bird.y >= HEIGHT - bird.radius:
            game_active = False
            save_score_to_excel()
    
    # Drawing
    if not name_input_active:
        for pipe in pipes:
            pipe.draw()
        
        bird.draw()
        draw_score()
        
        if not game_active and not name_input_active:
            draw_game_over()
    
    pygame.display.update()

pygame.quit()
sys.exit()